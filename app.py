import os
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.utils import secure_filename
from io import BytesIO

app = Flask(__name__)

# Configuration
MESSAGES_FILE = os.path.join(os.path.dirname(__file__), 'messages.xlsx')
HERO_FOLDER = os.path.join(app.static_folder, 'images', 'hero')
TEAM_FOLDER = os.path.join(app.static_folder, 'images', 'team')
LEGAL_FOLDER = os.path.join(app.static_folder, 'images', 'legal')
VISIMISI_FOLDER = os.path.join(app.static_folder, 'images', 'visimisi')
ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'gif', 'webp'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

# Create required folders if not exists
os.makedirs(HERO_FOLDER, exist_ok=True)
os.makedirs(TEAM_FOLDER, exist_ok=True)
os.makedirs(LEGAL_FOLDER, exist_ok=True)
os.makedirs(VISIMISI_FOLDER, exist_ok=True)

def init_excel():
    """Initialize Excel file with headers if it doesn't exist"""
    if not os.path.exists(MESSAGES_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"
        
        # Add headers
        headers = ["No.", "Nama", "Email", "Pesan", "Tanggal & Waktu"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 20
        
        try:
            wb.save(MESSAGES_FILE)
        except PermissionError:
            raise PermissionError(f"Tidak dapat membuat/menulis file '{MESSAGES_FILE}'. Tutup file jika sedang dibuka dan pastikan aplikasi memiliki izin tulis.")

def save_message(name, email, message):
    """Save message to Excel file"""
    init_excel()
    
    try:
        wb = load_workbook(MESSAGES_FILE)
    except PermissionError:
        raise PermissionError(f"Tidak dapat membuka file '{MESSAGES_FILE}' untuk menulis. Tutup file jika sedang dibuka dan pastikan aplikasi memiliki izin tulis.")
    ws = wb.active
    
    # Get the next row number
    row_num = ws.max_row + 1
    
    # Add data
    ws[f'A{row_num}'] = row_num - 1
    ws[f'B{row_num}'] = name
    ws[f'C{row_num}'] = email
    ws[f'D{row_num}'] = message
    ws[f'E{row_num}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Center align the row number
    ws[f'A{row_num}'].alignment = Alignment(horizontal="center", vertical="center")
    
    try:
        wb.save(MESSAGES_FILE)
    except PermissionError:
        raise PermissionError(f"Tidak dapat menyimpan file '{MESSAGES_FILE}'. Tutup file jika sedang dibuka dan pastikan aplikasi memiliki izin tulis.")


def save_consultation(name, email, service, preferred, message):
    """Save consultation entry into a 'Consultations' sheet in the messages Excel file."""
    init_excel()

    try:
        wb = load_workbook(MESSAGES_FILE)
    except PermissionError:
        raise PermissionError(f"Tidak dapat membuka file '{MESSAGES_FILE}' untuk menulis. Tutup file jika sedang dibuka dan pastikan aplikasi memiliki izin tulis.")

    # Create 'Consultations' sheet if it doesn't exist
    if 'Consultations' not in wb.sheetnames:
        ws = wb.create_sheet('Consultations')
        headers = ["No.", "Nama", "Email", "Layanan", "Waktu Preferensi", "Pesan", "Tanggal & Waktu"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 20

    ws = wb['Consultations']

    row_num = ws.max_row + 1
    ws[f'A{row_num}'] = row_num - 1
    ws[f'B{row_num}'] = name
    ws[f'C{row_num}'] = email
    ws[f'D{row_num}'] = service
    ws[f'E{row_num}'] = preferred
    ws[f'F{row_num}'] = message
    ws[f'G{row_num}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws[f'A{row_num}'].alignment = Alignment(horizontal="center", vertical="center")

    try:
        wb.save(MESSAGES_FILE)
    except PermissionError:
        raise PermissionError(f"Tidak dapat menyimpan file '{MESSAGES_FILE}'. Tutup file jika sedang dibuka dan pastikan aplikasi memiliki izin tulis.")

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_hero_images():
    """Get all hero images from folder, sorted"""
    if os.path.isdir(HERO_FOLDER):
        images = sorted([f for f in os.listdir(HERO_FOLDER) if allowed_file(f)])
        return images
    return []

def get_hero_image():
    """Get current hero image from folder"""
    images = get_hero_images()
    if images:
        return images[-1]  # Return latest image
    return None

def get_visimisi_images():
    """Get all visimisi images from folder, sorted"""
    if os.path.isdir(VISIMISI_FOLDER):
        images = sorted([f for f in os.listdir(VISIMISI_FOLDER) if allowed_file(f)])
        return images
    return []

def get_visimisi_image():
    """Get current visimisi background image from folder"""
    images = get_visimisi_images()
    if images:
        return images[-1]  # Return latest image
    return None



def delete_old_hero_images():
    """Delete old hero images, keep only the latest"""
    if os.path.isdir(HERO_FOLDER):
        images = sorted([f for f in os.listdir(HERO_FOLDER) if allowed_file(f)])
        if len(images) > 1:
            # Delete all but the latest
            for old_image in images[:-1]:
                try:
                    os.remove(os.path.join(HERO_FOLDER, old_image))
                except:
                    pass


def get_team_images():
    """Get all team images from the team folder."""
    if os.path.isdir(TEAM_FOLDER):
        return sorted([f for f in os.listdir(TEAM_FOLDER) if allowed_file(f)])
    return []


def get_legal_images():
    """Get all legal images from the legal folder."""
    if os.path.isdir(LEGAL_FOLDER):
        return sorted([f for f in os.listdir(LEGAL_FOLDER) if allowed_file(f)])
    return []


@app.route("/")
def home():
    hero_image = get_hero_image()
    visimisi_image = get_visimisi_image()
    return render_template("pages/home.html", hero_image=hero_image, visimisi_image=visimisi_image)


@app.route("/about")
def about():
    team_images = get_team_images()
    legal_images = get_legal_images()
    return render_template("pages/about.html", team_images=team_images, legal_images=legal_images)


@app.route("/services")
def services():
    return redirect(url_for('home') + '#services')


@app.route("/gallery")
def gallery():
    image_folder = os.path.join(app.static_folder, 'images')
    valid_ext = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.svg'}
    images = []

    if os.path.isdir(image_folder):
        images = sorted([
            filename for filename in os.listdir(image_folder)
            if os.path.splitext(filename)[1].lower() in valid_ext
        ])

    return render_template("pages/gallery.html", images=images)


@app.route("/admin")
def admin():
    current_hero = get_hero_image()
    return render_template("pages/admin.html", current_hero_image=current_hero)


@app.route("/admin/upload-hero", methods=["POST"])
def upload_hero():
    """Handle hero image upload"""
    if 'heroImage' not in request.files:
        return jsonify({'success': False, 'error': 'Tidak ada file yang dipilih'}), 400
    
    file = request.files['heroImage']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Tidak ada file yang dipilih'}), 400
    
    # Check file size
    if len(file.read()) > MAX_FILE_SIZE:
        file.seek(0)
        return jsonify({'success': False, 'error': 'Ukuran file terlalu besar. Maksimal 5MB.'}), 400
    
    file.seek(0)
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Format file tidak didukung. Gunakan JPG, PNG, GIF, atau WEBP.'}), 400
    
    try:
        # Generate secure filename with timestamp
        filename = secure_filename(f"hero_{datetime.now().timestamp()}_{file.filename}")
        filepath = os.path.join(HERO_FOLDER, filename)
        file.save(filepath)
        
        # Delete old images
        delete_old_hero_images()
        
        return jsonify({'success': True, 'message': 'Gambar berhasil diupload!'}), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/contact")
def contact():
    return render_template("pages/contact.html")


@app.route("/consultation", methods=["GET", "POST"])
def consultation():
    if request.method == "POST":
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        service = request.form.get('service', '').strip()
        preferred = request.form.get('preferred', '').strip()
        message = request.form.get('message', '').strip()

        if not name or not email or not service:
            return jsonify({'success': False, 'error': 'Nama, email, dan layanan wajib diisi'}), 400

        try:
            save_consultation(name, email, service, preferred, message)
            return jsonify({'message': 'Permintaan konsultasi berhasil dikirim!', 'success': True}), 200
        except PermissionError as e:
            return jsonify({'success': False, 'error': str(e)}), 500
        except Exception:
            return jsonify({'success': False, 'error': 'Terjadi kesalahan saat menyimpan permintaan konsultasi.'}), 500

    return render_template("pages/consultation.html")


if __name__ == "__main__":
    app.run(debug=True)